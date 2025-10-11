"""
File:   export_xlsx.py 
Desc:   Script function to xlsx

Copyright (C) 2024 Peter Himmler
Apache License 2.0
"""

from il_conv import VERSION_STR

from datetime import datetime

import openpyxl
import openpyxl.workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


import openpyxl.styles
import openpyxl.worksheet.table as xltables

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.cell import Cell
from openpyxl.styles import Font, Alignment


from resources import LOGO_PNG
from issuedb import IssueDB
from parse import LogDB

from export import Formatmode

def _map_dtype_2_auto_judgement(d : str ):
    '''
        d/p: detection type encodes up to for information seperated by ';'
             d or p for "definite occureance" or "potential occureance"
        n/c: assembly comparison results
             n or c for "no change in assembly comparison" and "change in assembly comparison"    
    '''
    # t, a, dir, f1, f2 
    l1 = ([ s.strip() for s in d.split(";")] + ["'?'"]*5)[:5]
    
    t, a, dir, f1, f2  = l1
    
    auto_judge = {
        ('p', '-'): "Potential. Further manual investigation required.",
        ('p', 'n'): "Potential. Most likely a false positive, ignore.",
        ('p', 'c'): f"Potential. Further manual investigation required. Assembly files '{f1}', '{f2}' generated  in '{dir}'.",
        ('d', '-'): "Definite. Most likely impacted. Check if mitigation is applied.",
        ('d', 'n'): "Definite. Unclear definite detection. Executed assembly comparison haven't shown a change ?? Reach out to vendor support-",
        ('d', 'c'): f"Definite. Most likely impacted. Assembly files '{f1}', '{f2}' generated in '{dir}'. Check if mitigation is applied.",
    }
   
    # Default result
    res : str = "?!?"
    
    # Use the tuple (t, a) as a key to fetch the appropriate message
    res = auto_judge.get((t, a), res)
    
    return res
    




def _addOneSheet(
    wb: Workbook, db: IssueDB, log_db: LogDB, fm: Formatmode, verbose: bool = False
):

    # dict of filename to filepath mapping
    Hide = False
    Visible = True
    fn2fp = {}
    detection : str
    worksheet_name = None
    ws: Worksheet = None  # Worksheet

    if fm == Formatmode.COMPACT:
        worksheet_name = "Report compact"
        ws = wb.create_sheet(worksheet_name)

        # (fieldname, type, Autofit, visible, , autofit max chars)
        col_style = [
            ("File Name", "file", False, Visible, -1),
            ("File Path", "string", True, Hide, 40),
            ("Detector", "string", False, Visible, -1),
            ("Issue ID", "hyper", False, Visible, 10),
            ("SIL", "string", False, Visible, -1),
            ("Fixed Version", "string", False, Visible, -1),
            ("Summary", "string", False, Visible, 60),
            ("Description", "string", False, Hide, 60),
            ("Mitigation", "string", False, Hide, 60),
            ("Lines", "string", True, Visible, -1),
            ("Auto judgement", "string", True, Visible, -1),

        ]
        headings = [f for (f, _, _, _, _) in col_style]
        # Add headings
        ws.append(headings)

        curs = log_db.conn.execute(
            "select file, filepath, issueid, group_concat(line), group_concat(detectiontype) FROM Logs GROUP By filepath, issueid ORDER BY file, issueid"
        )

        for fn, fp, id, lines, detection in curs:

            if fn2fp.get(fn) is not None:
                (count, existing_fp) = fn2fp[fn]
                if existing_fp != fp:
                    print(
                        f"WARN: Your project seems to have multiple times file '{fn}' in different folders, you should use expanded format and looking at the full pathname within the report."
                    )
                    fn2fp[fn] = (count + 1, fp)
            else:
                # first time or reoccurance of same fp for a fn
                fn2fp[fn] = (1, fp)

            # print (row)
            ii = db.getIssue(id)

            assert (
                ii is not None
            ), f"ERROR: Log includes detected issue id '{id}' in file '{fp}' \
                 but we have no information about it. Are you a current issue portal XML export and Inspector release note?"

            auto_judgement = _map_dtype_2_auto_judgement(detection)

            csvrow = [
                fn,
                fp,
                ii.detectiontype,
                ii.id,
                ii.sil,
                ii.fix_version,
                ii.summary,
                ii.description,
                ii.mitigation,
                lines,
                auto_judgement,
                ]

            ws.append(csvrow)

    elif fm == Formatmode.NORMAL:
        worksheet_name = "Report normal"
        ws = wb.create_sheet(worksheet_name)

        # (fieldname, type, Autofit, visible, , autofit max chars)
        col_style = [
            ("File Name", "file", False, Visible, -1),
            ("File Path", "string", True, Hide, 40),
            ("Detector", "string", False, Visible, -1),
            ("Issue ID", "hyper", False, Visible, 10),
            ("SIL", "string", False, Visible, -1),
            ("Fixed Version", "string", False, Visible, -1),
            ("Summary", "string", False, Visible, 60),
            ("Description", "string", False, Hide, 0),
            ("Mitigation", "string", False, Hide, 60),
            ("Line", "int", True, Visible, -1),
            ("Column", "int", True, Visible, -1),
            ("Auto judgement", "string", True, Visible, -1),
            ("Resolved/Checked", "string", True, Visible, -1),
        ]
        headings = [f for (f, _, _, _, _) in col_style]

        # Add headings
        ws.append(headings)

        curs = log_db.conn.execute(
            "select file, filepath, issueid, line, column, detectiontype FROM Logs ORDER By file, rowid"
        )

        for fn, fp, id, line, column, detection in curs:

            if fn2fp.get(fn) is not None:
                (count, existing_fp) = fn2fp[fn]
                if existing_fp != fp:
                    print(
                        f"WARN: Your project seems to have multiple times file '{fn}' in different folders, you should use expanded format and looking at the full pathname within the report."
                    )
                    fn2fp[fn] = (count + 1, fp)
            else:
                # first time or reoccurance of same fp for a fn
                fn2fp[fn] = (1, fp)

            ii = db.getIssue(id)

            assert (
                ii is not None
            ), f"ERRRO: Log includes detected issue id but we have no information about it.\n{id} {fp} line"

            auto_judgement = _map_dtype_2_auto_judgement(detection)
             
            csvrow = [
                fn,
                fp,
                ii.detectiontype,
                ii.id,
                ii.sil,
                ii.fix_version,
                ii.summary,
                ii.description,
                ii.mitigation,
                line,
                column,
                auto_judgement,
                "not checked",
            ]

            ws.append(csvrow)

    elif fm == Formatmode.EXTENDED:
        worksheet_name = "Report extended"
        ws = wb.create_sheet(worksheet_name)

        # (fieldname, type, Autofit, visible, autofit max chars)
        col_style = [
            ("File Name", "file", False, Visible, -1),
            ("File Path", "string", True, Visible, 40),
            ("Detector", "string", False, Visible, -1),
            ("Issue ID", "hyper", False, Visible, 10),
            ("SIL", "string", False, Visible, -1),
            ("Fixed Version", "string", False, Visible, -1),
            ("Summary", "string", False, Visible, 60),
            ("Description", "string", False, Visible, 70),
            ("Mitigation", "string", False, Visible, 70),
            ("Line", "int", True, Visible, -1),
            ("Column", "int", True, Visible, -1),
            ("Auto judgement", "string", True, Visible, -1),
            ("Resolved/Checked", "string", True, Visible, -1),
        ]
        headings = [f for (f, _, _, _, _) in col_style]

        # Add headings
        ws.append(headings)

        curs = log_db.conn.execute(
            "select file, filepath, issueid, line, column, detectiontype FROM Logs ORDER By file, issueid, line"
        )

        for fn, fp, id, line, column, detection in curs:
            if fn2fp.get(fn) is not None:
                (count, existing_fp) = fn2fp[fn]
                if existing_fp != fp:
                    print(
                        f"WARN: Your project seems to have multiple times file '{fn}' in different folders, you should use expanded format and looking at the full pathname within the report."
                    )
                    fn2fp[fn] = (count + 1, fp)
            else:
                # first time or reoccurance of same fp for a fn
                fn2fp[fn] = (1, fp)

            ii = db.getIssue(id)

            assert (
                ii is not None
            ), f"ERRRO: Log includes detected issue id but we have no information about it.\n{id} {fp} line"

            auto_judgement = _map_dtype_2_auto_judgement(detection)

            csvrow = [
                fn,
                fp,
                ii.detectiontype,
                ii.id,
                ii.sil,
                ii.fix_version,
                ii.summary,
                ii.description,
                ii.mitigation,
                line,
                column,
                auto_judgement,
                "not checked",
            ]

            ws.append(csvrow)

    # calculate max # of character for all columns
    max_chars = []
    for col in ws.iter_cols(min_row=1, max_col=ws.max_column):
        max_chars.append(max(len(str(cell.value)) for cell in col))

    # define dimension and style of column
    dim_holder = ws.column_dimensions
    for i, col in enumerate(ws.iter_cols(min_row=1, max_col=ws.max_column, max_row=1)):
        #        for column_cells in ws['A:H']:
        _, _, _, _, pro_chars = col_style[i]
        column_letter = get_column_letter(col[0].column)

        if pro_chars == -1:
            pro_chars = max_chars[i]

        #        ws.column_dimensions[column_letter].hidden = not visible
        dim_holder[column_letter].width = min(100, (pro_chars + 2) * 1.23)

    # Add one addtional row which will hold our logo and title :-)
    ws.insert_rows(1)
    
    # Future table starts at column A
    dim_holder = ws.column_dimensions
    fieldname, types, autofits, visiblities, maxchars = list(zip(*col_style))
    
    # Iterate through all columns (col_style)
    for i in range(0, len(types)):
        type_str = types[i]
        column_letter = get_column_letter(i + 1)  # i+1 as Excel columns start with A = 1)

        dim_holder[column_letter].hidden = not visiblities[i]

        # iterate all cells with the current referenced column row 3 ... to max
        for cell in ws[column_letter][2:]:    # skip our title row (row 1) and our future table header row (row 2)
        
            cell.alignment = Alignment(wrap_text=autofits[i])
            cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT
        
            # check for special extra markup for the cell based on our specification col_style
            if "file" in type_str:
                try:
                    (count, fp) = fn2fp[cell.value]
                    
                    # cross-check: filename <-> fullpath mapping and if there are multiplte filenames in the project in different paths
                    if count > 1 and fm == Formatmode.COMPRESSED:
                        cell.comment = openpyxl.comments.Comment(
                            "HINT: File name '{}' is not unique within your project - the compressed formatting report might wrongly mix multiple occurances!!\n{}".format(
                                cell.value, fp
                            ),
                            "generated",
                            100,
                            640,
                        )
                except:
                    pass
                
            elif "datetime" in type_str:
                try:
                    # type str here assumes "datetime;FORMATSTRING"
                    cell.value = datetime.strptime(cell.value, type_str.split(";")[-1])
                except:
                    pass
                cell.number_format = openpyxl.styles.numbers.FORMAT_DATE_DDMMYY

            elif "int" in type_str:
                try:
                    cell.value = int(cell.value)
                    cell.number_format = openpyxl.styles.numbers.FORMAT_NUMBER
                except:
                    pass
                
            elif "hyper" in type_str:
                try:
                    ##
                    ## NOTE: A more reliable approach might be to convert the value into formulat
                    ##       "=HYPERLINK(....)" but did not check that
                    ##
                    ## current apporach is to use the cell styling, but looks that the library or excel has a bug
                    ## After styling no ws changes like insert/remove a row is allowed as otherwise hyperlink is attached
                    ## to the wrong cell after loading in EXCEL. OMG - GRRRRR 
                    ##
                    
                    cell.style = "Hyperlink"

                    id = str(cell.value)  # .split('-')[-1]
                    if id.startswith("TCVX-") or id.startswith("SMRT-"):
                        ii = db.getIssue(id)
                        cell.comment = openpyxl.comments.Comment(
                            "MITIGATION:\n{}".format(ii.mitigation),
                            "generated",
                            400,
                            520,
                        )
                        uri = f"https://issues.tasking.com/?issueid={id}"
                        cell.hyperlink = uri
                except:
                    pass
    
    # define the table at A2 
    cell_range = CellRange( min_col=1, min_row=2, max_col=ws.max_column, max_row=ws.max_row )

    table_name = "Data_" + str(fm)[str(fm).find(".") + 1 :]
    tab = xltables.Table(displayName=table_name, ref=str(cell_range))
    tab.tableStyleInfo = xltables.TableStyleInfo(name="TableStyleLight9")
    ws.add_table(tab)


    img = openpyxl.drawing.image.Image(LOGO_PNG)
    img.anchor = "A1"
    img.width = 48
    img.height = 48
    ws.add_image(img)
    openpyxl.comments.Comment

    ws.row_dimensions[1].height = 40
    ws["D1"].alignment = Alignment(vertical="center")
    ws["D1"] = worksheet_name
    ws["D1"].font = Font(bold=True, size=26)

    if verbose:
        print(f"INFO:  Generating Excel Worksheet '{worksheet_name}'")


def generateExcel(
    output_file_name: str, db: IssueDB, log_db: LogDB, verbose: bool = False
):
    """Generate Excel output.

    Args:
        output_file_name (str): The name of the file to save
        db (IssueDB): IssueDB from portal XML export related with Inspector release note
        log_db (LogDB): Database from parse log detection entries
        fm (FormatMode): Enum value to configure generator.
        verbose (bool): Create verbose output during processing
    """

    if verbose:
        print("INFO: Generating Excel Workbook")

    wb = Workbook()
    wb.iso_dates = True
    ws = wb.active
    ws.title = "TriCore Inspector Reports"

    img = openpyxl.drawing.image.Image(LOGO_PNG)
    img.anchor = "A1"
    img.width = 64
    img.height = 64
    ws.add_image(img)
    openpyxl.comments.Comment

    ws.row_dimensions[1].height = 54
    ws["B1"].alignment = Alignment(vertical="center")
    ws["B1"] = ws.title
    ws["B1"].font = Font(bold=True, size=26)

    ws["A3"] = "Generated at:"
    ws["A3"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["A3"].font = Font(bold=True, size=12)

    ws["B3"] = datetime.today().isoformat()
    ws["B3"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["B3"].font = Font(size=12)

    ws["A4"] = "Generated by:"
    ws["A4"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["A4"].font = Font(bold=True, size=12)

    ws["B4"] = f"il_conv ({VERSION_STR})"
    ws["B4"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["B4"].font = Font(size=12)

    ws["A5"] = "SPDX short identifier:"
    ws["A5"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["A5"].font = Font(bold=True, size=12)

    ws["B5"] = "Apache-2.0"
    ws["B5"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["B5"].font = Font(size=12)

    ws["A6"] = "Repository:"
    ws["A6"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["A6"].font = Font(bold=True, size=12)
    ws["B6"] = "https://github.com/Paul-Hi/il_conv"
    ws["B6"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["B6"].font = Font(size=12)

    ws["A8"] = "XML data source:"
    ws["A8"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["A8"].font = Font(bold=True, size=12)
    ws["B8"] = str(db.xmlfile)
    ws["B8"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["B8"].font = Font(size=12)

    ws["A9"] = "Inspector data source:"
    ws["A9"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["A9"].font = Font(bold=True, size=12)
    ws["B9"] = str(db.relnotefile)
    ws["B9"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["B9"].font = Font(size=12)

    ws["A10"] = "Compiler:"
    ws["A10"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["A10"].font = Font(bold=True, size=12)
    ws["B10"] = db.compiler_version
    ws["B10"].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    ws["B10"].font = Font(size=12)

    # extend width
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    _addOneSheet(wb, db, log_db, Formatmode.COMPACT, verbose)
    _addOneSheet(wb, db, log_db, Formatmode.NORMAL, verbose)
    _addOneSheet(wb, db, log_db, Formatmode.EXTENDED, verbose)

    if verbose:
        print(f"INFO: Written to file '{output_file_name}'")

    wb.save(filename=output_file_name)
