"""
File:   export.py
Desc:   Exporter classes to export the generated data to xlsx or html.

Copyright (C) 2022 Paul Himmler, Peter Himmler
Apache License 2.0
"""

from datetime import datetime
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.styles import Font
import openpyxl.worksheet.table as xltables
from resources import resource_path
from issue_store import Portalissue
import itertools
import tables
import parse
import configs


class Exporter(object):
    '''The Exporter is an interface for all other exporters.'''

    def export(self, table: tables.Table, file_name: str, configuration: configs.Config, verbose: bool, db: dict = {}):
        '''Exports a table to some output format

        Args:
            table(Table): The table to export
            file_name (str): The name of the file to put out without the extension
            verbose (bool): Create verbose output during processing
            configuration(config): Table configuration
            db(dict): Additional information on issues - if any

        '''

        pass


class ExcelExporter(Exporter):
    '''The ExcelExporter is an exporter for exports to xlsx files.'''

    def export(self, table: tables.Table, file_name: str, configuration: configs.Config, verbose: bool, db: dict = {}):
        '''Exports a table formatted to a xlsx file

        Args:
            table(Table): The table to export
            file_name (str): The name of the file to put out without the extension
            configuration(config): Table configuration
            verbose (bool): Create verbose output during processing
            db(dict): Additional information on issues - if any

        '''

        if verbose:
            print(".... Creating Workbook")

        wb = Workbook()
        wb.iso_dates = True
        ws = wb.active
        ws.title = "INSPECTOR Msg"

        # note we insert the first line at the end now - makes the table creation easier

        headings = configuration.get_column_names()
        types = configuration.get_column_types()

        # Add additional columns from Issue Portal
        ws.append(headings)

        all_rows = list(itertools.chain(*[page.rows for page in table.pages]))
        i = 0
        for row in all_rows:
            i = i+1
            # lookup issue id within portal database
            ticket_id = row.get(parse.TICKET_ID_KEY)
            issue_entry = db.get(ticket_id, Portalissue())._asdict()
            issue_columns = dict((key, entry) for (
                key, entry) in issue_entry.items() if key in configuration.get_column_names())

            # add all diagnostic message columns + all information from issue portal
            row_or_default = [row[key] if key in row else issue_columns.get(
                key, "") for key in configuration.get_column_names()]
            ws.append(row_or_default)

        cell_range = CellRange(min_col=1, min_row=2, max_col=len(
            headings), max_row=table.row_count + 2)

        dim_holder = ws.column_dimensions
        for column_cells in ws['A:Z']:
            max_chars = max(len(str(cell.value)) for cell in column_cells)
            column_letter = (get_column_letter(column_cells[0].column))
            if max_chars > 0:
                dim_holder[column_letter].width = min(
                    80, (max_chars + 2) * 1.15)

        # we know the table starts at column A, so we can use the types
        # types[0] is type for column A and so on ...
        for i, type_str in enumerate(types):
            if "datetime" in type_str:
                column_letter = get_column_letter(i + 1)
                for cell in ws[column_letter][1:]:
                    try:
                        cell.value = datetime.strptime(
                            cell.value, type_str.split(';')[-1])
                    except:
                        pass
                    cell.number_format = "YYYY-MM-DD HH:MM:SS"
            elif "int" in type_str:
                column_letter = get_column_letter(i + 1)
                for cell in ws[column_letter][1:]:
                    try:
                        cell.value = int(cell.value)
                    except:
                        pass
                    cell.number_format = "0"

        tab = xltables.Table(displayName="Data", ref=str(cell_range))

        style = xltables.TableStyleInfo(name="TableStyleLight9")

        tab.tableStyleInfo = style
        ws.add_table(tab)

        ws.insert_rows(1)

        ws['B1'] = 'Tricore Inspector Diagnostic message output.'
        img = openpyxl.drawing.image.Image(resource_path("res/logo.png"))
        img.anchor = 'A1'
        img.width = 64
        img.height = 64
        ws.add_image(img)
        ws['A1'].style
        # change font for first line
        for row_cells in ws.iter_rows(min_row=1, max_row=1, min_col=2, max_col=2):
            for c in row_cells:
                c.font = Font(bold=True, size=38)

        wb.save(filename=file_name+".xlsx")


class HTMLExporter(Exporter):
    '''The HTMLExporter is an exporter for exports to html files.'''

    def export(self, table: tables.Table, file_name: str, configuration: configs.Config, verbose: bool, db: dict = {}):
        '''Exports a table formatted to a htnl file

        Args:
            table(Table): The table to export
            file_name (str): The name of the file to put out without the extension
            configuration(config): Table configuration
            verbose (bool): Create verbose output during processing
            db(dict): Additional information on issues - if any

        '''

        if verbose:
            print(".... Creating HTML output")

        with open(resource_path("res/default.css")) as css:
            css_style = css.read()
        with open(resource_path("res/functions.js")) as functions:
            js_functions = functions.read()
        with open(resource_path("res/logo_base64.txt")) as img:
            image_b64 = img.read().strip().replace('\n', '')

        html = '<!DOCTYPE html>'
        html += '\n<html>'
        html += '\n<head>'
        # html += '\n<link rel="stylesheet" href="default.css">'
        html += '\n<style>\n' + css_style + '\n</style>'
        html += '\n<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/4.7.0/css/font-awesome.min.css">'
        # html += '\n<script src="functions.js"></script>'
        html += '\n<script>\n' + js_functions + '\n</script>'
        html += '\n</head>'
        html += '\n<body>'

        # logo_img = '\n<img src="logo.png" id="Logo" alt="IL Converter Logo" height="96px" width=auto>'
        logo_img = '\n <img src="data:image/png;base64,' + image_b64 + \
            '" id="Logo" alt="IL Converter Logo" height="96px" width=auto>'

        html += logo_img

        order_arrows = '\n     <i class="fa fa-caret-up" aria-hidden="true"></i>\n     <i class="fa fa-caret-down" aria-hidden="true"></i>'

        # extend headings and types with issue portal data
        headings = configuration.get_column_names()
        types = configuration.get_column_types()
        assert(len(types) == len(headings))

        # get rid of format stuff
        types = [tp.split(';')[0] for tp in types]

        html_table = '\n  <thead>'
        html_table += '\n   <tr>' + \
            ''.join('\n    <th class="draggable" draggable="true" data-type="' +
                    types[i] + '">' + h + order_arrows + '\n    </th>' for i, h in enumerate(headings)) + '\n   </tr>'
        html_table += '\n  </thead>'

        html_table += '\n  <tbody>'
        all_rows = list(itertools.chain(*[page.rows for page in table.pages]))
        for row in all_rows:
            # lookup issue id within portal database
            ticket_id = row.get(parse.TICKET_ID_KEY)
            issue_entry = db.get(ticket_id, Portalissue())._asdict()
            issue_columns = dict((key, entry) for (
                key, entry) in issue_entry.items() if key in configuration.get_column_names())

            # add all diagnostic message columns + all information from issue portal
            row_or_default = [row[key] if key in row else issue_columns.get(
                key, "") for key in configuration.get_column_names()]
            html_table += '\n <tr>\n'
            html_table += '\n'.join('  <td>' + str(entry) +
                                    '</td>' for entry in row_or_default)
            html_table += '\n </tr>'

        html_table += '\n</tbody>\n'

        html_table = '\n <table border=1 class="simple_table" id="log_table">' + \
            html_table + '\n </table>'

        html += html_table

        html += '\n</body>'
        html += '\n</html>'

        with open(file_name + ".html", "w") as output_file:
            output_file.write(html)
