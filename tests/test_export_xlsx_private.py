"""
File:   test_export_xlsx_private.py
Desc:   Integration tests for export_xlsx.py using local (non-committed) XML and
        release note files from the XML/ and RELEASENOTES/ folders.

        Tests are skipped automatically when the required files are absent,
        so this file is safe to run on any machine.

Copyright (C) 2024 Peter Himmler
Apache License 2.0
"""

import gc
import io
import os
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook

from export import Formatmode
from export_xlsx import _addOneSheet, generateExcel
from issuedb import IssueDB
from parse import LogDB

# ---------------------------------------------------------------------------
# Paths to local test data (not committed)
# ---------------------------------------------------------------------------

_HERE = Path(__file__).parent.parent  # project root

XML_V63R1   = _HERE / "XML"          / "issues_tasking_TCVX_v6.3r1.xml"
RN_V63_V108 = _HERE / "RELEASENOTES" / "readme_tricore_v6.3r1_inspector_v1.0r8.html"

_HAVE_DATA = XML_V63R1.exists() and RN_V63_V108.exists()
_SKIP = unittest.skipUnless(_HAVE_DATA, "Local v6.3r1 / v1.0r8 test data not available")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_tmp(content: str) -> str:
    f = tempfile.NamedTemporaryFile(mode="w", suffix=".log", delete=False)
    f.write(content)
    f.close()
    return f.name


def _close_db(db: IssueDB):
    """Explicitly close the SQLite connection — required on Windows before unlink."""
    if db.conn:
        db.conn.close()
        db.conn = None


def _remove_db(compiler_ver: str, inspector_ver: str):
    try:
        Path("issues-{}-{}.db".format(compiler_ver, inspector_ver)).unlink()
    except FileNotFoundError:
        pass


def _xlsx_tmp() -> str:
    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    f.close()
    return f.name


def _capture(fn, *args, **kwargs):
    """Call fn(*args, **kwargs), capture and return stdout as a string."""
    buf = io.StringIO()
    sys.stdout = buf
    try:
        fn(*args, **kwargs)
    finally:
        sys.stdout = sys.__stdout__
    return buf.getvalue()


def _col_index(ws, header_row: int, name: str) -> int:
    """Return 0-based column index of *name* in the given header row."""
    for i, cell in enumerate(ws[header_row]):
        if cell.value == name:
            return i
    raise KeyError(f"Column '{name}' not found in row {header_row}")


# Issue IDs present in both v6.3r1 XML and v1.0r8 release note
_ID_POTENTIAL    = "TCVX-39753"   # Potential / insp_ctc
_ID_DEFINITE     = "TCVX-39025"   # Definite  / insp_ltc
_ID_ASM_NOCHANGE = "TCVX-40928"   # Potential / insp_ltc — used for W981 (p;n)
_ID_ASM_CHANGE   = "TCVX-43123"   # Definite  / insp_astc — used for E982 (p;c)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

@_SKIP
class TestGenerateExcelBasic(unittest.TestCase):
    """generateExcel happy path: workbook structure, cover sheet, report sheets."""

    @classmethod
    def setUpClass(cls):
        cls.db = IssueDB("v6.3r1", "v1.0r8", XML_V63R1, RN_V63_V108, verbose=False)
        cls.db.import_release_note()
        cls.db.import_xml_file()

        log_content = (
            f'W998: ["C:/src/foo.c" 10/1] [INSP] detected potential occurrence of issue {_ID_POTENTIAL}.\n'
            f'W999: ["C:/src/bar.c" 20/2] [INSP] detected occurrence of issue {_ID_DEFINITE}.\n'
        )
        cls._log = _write_tmp(log_content)
        cls.log_db = LogDB()
        cls.log_db.parse_log_file(cls._log)

        cls._xlsx = _xlsx_tmp()
        generateExcel(cls._xlsx, cls.db, cls.log_db, verbose=False)
        cls.wb = openpyxl.load_workbook(cls._xlsx)

    @classmethod
    def tearDownClass(cls):
        _close_db(cls.db)
        del cls.db
        gc.collect()
        _remove_db("v6.3r1", "v1.0r8")
        os.unlink(cls._log)
        os.unlink(cls._xlsx)

    # --- workbook ---

    def test_expected_sheet_names(self):
        self.assertEqual(
            self.wb.sheetnames,
            ["TriCore Inspector Reports", "Report compact", "Report extended"],
        )

    # --- cover sheet ---

    def test_cover_title_cell(self):
        self.assertEqual(self.wb["TriCore Inspector Reports"]["B1"].value, "TriCore Inspector Reports")

    def test_cover_generated_by_contains_il_conv(self):
        self.assertIn("il_conv", self.wb["TriCore Inspector Reports"]["B4"].value)

    def test_cover_xml_source_label(self):
        self.assertEqual(self.wb["TriCore Inspector Reports"]["A8"].value, "XML data source:")

    def test_cover_xml_source_path(self):
        self.assertIn("v6.3r1", self.wb["TriCore Inspector Reports"]["B8"].value)

    def test_cover_compiler_version(self):
        self.assertEqual(self.wb["TriCore Inspector Reports"]["B10"].value, "v6.3r1")

    # --- compact sheet structure ---

    def test_compact_title_row(self):
        ws = self.wb["Report compact"]
        self.assertEqual(ws["D1"].value, "Report compact")

    def test_compact_headers(self):
        ws = self.wb["Report compact"]
        headers = [c.value for c in ws[2]]
        self.assertIn("File Name", headers)
        self.assertIn("Issue ID", headers)
        self.assertIn("SIL", headers)
        self.assertIn("Auto judgement", headers)
        self.assertNotIn("Column", headers)   # COMPACT has no Column field

    def test_compact_has_data_rows(self):
        ws = self.wb["Report compact"]
        data_rows = [r for r in ws.iter_rows(min_row=3, values_only=True) if any(v is not None for v in r)]
        self.assertGreater(len(data_rows), 0)

    def test_compact_contains_potential_issue(self):
        ws = self.wb["Report compact"]
        col = _col_index(ws, 2, "Issue ID")
        ids = [ws.cell(row=r, column=col + 1).value for r in range(3, ws.max_row + 1)]
        self.assertIn(_ID_POTENTIAL, ids)

    def test_compact_contains_definite_issue(self):
        ws = self.wb["Report compact"]
        col = _col_index(ws, 2, "Issue ID")
        ids = [ws.cell(row=r, column=col + 1).value for r in range(3, ws.max_row + 1)]
        self.assertIn(_ID_DEFINITE, ids)

    def test_compact_auto_judgement_potential(self):
        ws = self.wb["Report compact"]
        id_col = _col_index(ws, 2, "Issue ID")
        aj_col = _col_index(ws, 2, "Auto judgement")
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=id_col + 1).value == _ID_POTENTIAL:
                self.assertIn("Potential", ws.cell(row=row, column=aj_col + 1).value)
                break

    def test_compact_auto_judgement_definite(self):
        ws = self.wb["Report compact"]
        id_col = _col_index(ws, 2, "Issue ID")
        aj_col = _col_index(ws, 2, "Auto judgement")
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=id_col + 1).value == _ID_DEFINITE:
                self.assertIn("Definite", ws.cell(row=row, column=aj_col + 1).value)
                break

    # --- extended sheet structure ---

    def test_extended_title_row(self):
        ws = self.wb["Report extended"]
        self.assertEqual(ws["D1"].value, "Report extended")

    def test_extended_headers(self):
        ws = self.wb["Report extended"]
        headers = [c.value for c in ws[2]]
        self.assertIn("Line", headers)
        self.assertIn("Column", headers)
        self.assertIn("Resolved/Checked", headers)

    def test_extended_has_data_rows(self):
        ws = self.wb["Report extended"]
        data_rows = [r for r in ws.iter_rows(min_row=3, values_only=True) if any(v is not None for v in r)]
        self.assertGreater(len(data_rows), 0)

    def test_extended_line_column_are_integers(self):
        ws = self.wb["Report extended"]
        line_col = _col_index(ws, 2, "Line")
        col_col  = _col_index(ws, 2, "Column")
        row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        self.assertIsInstance(row3[line_col], int)
        self.assertIsInstance(row3[col_col], int)

    def test_extended_resolved_column_value(self):
        ws = self.wb["Report extended"]
        rc_col = _col_index(ws, 2, "Resolved/Checked")
        row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        self.assertEqual(row3[rc_col], "not checked")


@_SKIP
class TestGenerateExcelVerbose(unittest.TestCase):
    """generateExcel verbose=True prints INFO lines."""

    @classmethod
    def setUpClass(cls):
        cls.db = IssueDB("v6.3r1", "v1.0r8", XML_V63R1, RN_V63_V108, verbose=False)
        cls.db.import_release_note()
        cls.db.import_xml_file()
        cls._log = _write_tmp(
            f'W998: ["C:/src/foo.c" 1/1] [INSP] detected potential occurrence of issue {_ID_POTENTIAL}.\n'
        )
        cls.log_db = LogDB()
        cls.log_db.parse_log_file(cls._log)

    @classmethod
    def tearDownClass(cls):
        _close_db(cls.db)
        del cls.db
        gc.collect()
        _remove_db("v6.3r1", "v1.0r8")
        os.unlink(cls._log)

    def test_verbose_prints_info(self):
        xlsx = _xlsx_tmp()
        try:
            output = _capture(generateExcel, xlsx, self.db, self.log_db, verbose=True)
            self.assertIn("INFO", output)
        finally:
            os.unlink(xlsx)


@_SKIP
class TestGenerateExcelAsmDetections(unittest.TestCase):
    """Assembly comparison detectiontypes (p;n and p;c) produce correct auto-judgements."""

    @classmethod
    def setUpClass(cls):
        cls.db = IssueDB("v6.3r1", "v1.0r8", XML_V63R1, RN_V63_V108, verbose=False)
        cls.db.import_release_note()
        cls.db.import_xml_file()

        log_content = (
            f'\tW981: ["C:/src/qux.c" 5/1] [INSP] detected potential occurrence of issue {_ID_ASM_NOCHANGE}. No change in assembly comparison detected.\n'
            f'E982: ["C:/src/quux.c" 8/2] [INSP] detected potential occurrence of issue {_ID_ASM_CHANGE}. Assembly files are stored in directory /tmp/asm as: affected.s; with fix: unaffected.s.\n'
        )
        cls._log = _write_tmp(log_content)
        cls.log_db = LogDB()
        cls.log_db.parse_log_file(cls._log)

        cls._xlsx = _xlsx_tmp()
        generateExcel(cls._xlsx, cls.db, cls.log_db, verbose=False)
        cls.wb = openpyxl.load_workbook(cls._xlsx)

    @classmethod
    def tearDownClass(cls):
        _close_db(cls.db)
        del cls.db
        gc.collect()
        _remove_db("v6.3r1", "v1.0r8")
        os.unlink(cls._log)
        os.unlink(cls._xlsx)

    def _aj_for_id(self, issue_id: str) -> str:
        ws = self.wb["Report extended"]
        id_col = _col_index(ws, 2, "Issue ID")
        aj_col = _col_index(ws, 2, "Auto judgement")
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=id_col + 1).value == issue_id:
                return ws.cell(row=row, column=aj_col + 1).value
        return None

    def test_p_n_produces_false_positive_judgement(self):
        aj = self._aj_for_id(_ID_ASM_NOCHANGE)
        self.assertIsNotNone(aj)
        self.assertIn("false positive", aj)

    def test_p_c_with_directory_produces_assembly_files_judgement(self):
        aj = self._aj_for_id(_ID_ASM_CHANGE)
        self.assertIsNotNone(aj)
        self.assertIn("Potential", aj)
        self.assertIn("affected.s", aj)
        self.assertIn("unaffected.s", aj)


@_SKIP
class TestGenerateExcelDuplicateFilename(unittest.TestCase):
    """Two detections for same filename from different paths trigger WARN in compact mode."""

    @classmethod
    def setUpClass(cls):
        cls.db = IssueDB("v6.3r1", "v1.0r8", XML_V63R1, RN_V63_V108, verbose=False)
        cls.db.import_release_note()
        cls.db.import_xml_file()

        log_content = (
            f'W998: ["C:/module_a/common.c" 5/1] [INSP] detected potential occurrence of issue {_ID_POTENTIAL}.\n'
            f'W998: ["C:/module_b/common.c" 15/2] [INSP] detected potential occurrence of issue {_ID_POTENTIAL}.\n'
        )
        cls._log = _write_tmp(log_content)
        cls.log_db = LogDB()
        cls.log_db.parse_log_file(cls._log)

    @classmethod
    def tearDownClass(cls):
        _close_db(cls.db)
        del cls.db
        gc.collect()
        _remove_db("v6.3r1", "v1.0r8")
        os.unlink(cls._log)

    def test_duplicate_filename_triggers_warn(self):
        xlsx = _xlsx_tmp()
        try:
            output = _capture(generateExcel, xlsx, self.db, self.log_db, verbose=False)
            self.assertIn("WARN", output)
        finally:
            os.unlink(xlsx)


@_SKIP
class TestAddOneSheetErrors(unittest.TestCase):
    """_addOneSheet raises ValueError when a log issue ID is absent from IssueDB."""

    @classmethod
    def setUpClass(cls):
        cls.db = IssueDB("v6.3r1", "v1.0r8", XML_V63R1, RN_V63_V108, verbose=False)
        cls.db.import_release_note()
        cls.db.import_xml_file()

    @classmethod
    def tearDownClass(cls):
        _close_db(cls.db)
        del cls.db
        gc.collect()
        _remove_db("v6.3r1", "v1.0r8")

    def _log_db_with(self, line: str) -> tuple[LogDB, str]:
        path = _write_tmp(line)
        db = LogDB()
        db.parse_log_file(path)
        return db, path

    def test_compact_raises_for_unknown_issue_id(self):
        log_db, path = self._log_db_with(
            'W998: ["C:/src/x.c" 1/1] [INSP] detected potential occurrence of issue TCVX-00000.\n'
        )
        try:
            with self.assertRaises(ValueError):
                _addOneSheet(Workbook(), self.db, log_db, Formatmode.COMPACT)
        finally:
            os.unlink(path)

    def test_extended_raises_for_unknown_issue_id(self):
        log_db, path = self._log_db_with(
            'W999: ["C:/src/y.c" 2/2] [INSP] detected occurrence of issue TCVX-00000.\n'
        )
        try:
            with self.assertRaises(ValueError):
                _addOneSheet(Workbook(), self.db, log_db, Formatmode.EXTENDED)
        finally:
            os.unlink(path)


if __name__ == "__main__":
    unittest.main(verbosity=2)
