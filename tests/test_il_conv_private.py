"""
File:   test_il_conv_private.py
Desc:   Integration tests for il_conv.py (the CLI entry point) using local
        (non-committed) XML and release note files and the committed test.log.

        Tests are skipped automatically when the required files are absent,
        so this file is safe to run on any machine.

Copyright (C) 2024 Peter Himmler
Apache License 2.0
"""

import gc
import io
import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

from il_conv import il_conv

# ---------------------------------------------------------------------------
# Paths to local test data
# ---------------------------------------------------------------------------

_HERE = Path(__file__).parent.parent  # project root

XML_V63R1   = _HERE / "XML"          / "issues_tasking_TCVX_v6.3r1.xml"
RN_V63_V108 = _HERE / "RELEASENOTES" / "readme_tricore_v6.3r1_inspector_v1.0r8.html"
TEST_LOG    = _HERE / "test.log"

_HAVE_DATA = XML_V63R1.exists() and RN_V63_V108.exists() and TEST_LOG.exists()
_SKIP = unittest.skipUnless(_HAVE_DATA, "Local v6.3r1 / v1.0r8 test data or test.log not available")

# il_conv derives inspector_version = "v1.0" from the release note filename
_DB_FILE = "issues-v6.3r1-v1.0.db"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _remove_db():
    try:
        Path(_DB_FILE).unlink()
    except FileNotFoundError:
        pass


def _run_il_conv(extra_args: list, capture_stdout: bool = False):
    """Call il_conv() with patched sys.argv.  Returns captured stdout or ''."""
    argv = [
        "il_conv.py",
        "-x", str(XML_V63R1),
        "-r", str(RN_V63_V108),
    ] + extra_args

    if not capture_stdout:
        with patch("sys.argv", argv):
            il_conv()
        return ""

    buf = io.StringIO()
    with patch("sys.argv", argv):
        sys.stdout = buf
        try:
            il_conv()
        finally:
            sys.stdout = sys.__stdout__
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

@_SKIP
class TestIlConvBasic(unittest.TestCase):
    """End-to-end run: output workbook is created with correct structure."""

    @classmethod
    def setUpClass(cls):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        cls._xlsx = tmp.name
        cls._stem = tmp.name[:-5]   # il_conv appends the extension itself

        _run_il_conv(["--output", cls._stem, str(TEST_LOG)])
        cls.wb = openpyxl.load_workbook(cls._xlsx)

    @classmethod
    def tearDownClass(cls):
        try:
            os.unlink(cls._xlsx)
        except FileNotFoundError:
            pass
        _remove_db()
        gc.collect()

    def test_sheet_names(self):
        self.assertEqual(
            self.wb.sheetnames,
            ["TriCore Inspector Reports", "Report compact", "Report extended"],
        )

    def test_cover_compiler_version(self):
        self.assertEqual(
            self.wb["TriCore Inspector Reports"]["B10"].value, "v6.3r1"
        )

    def test_cover_xml_source_label(self):
        self.assertEqual(
            self.wb["TriCore Inspector Reports"]["A8"].value, "XML data source:"
        )

    def test_compact_has_data_rows(self):
        ws = self.wb["Report compact"]
        rows = [r for r in ws.iter_rows(min_row=3, values_only=True)
                if any(v is not None for v in r)]
        self.assertGreater(len(rows), 0)

    def test_extended_has_data_rows(self):
        ws = self.wb["Report extended"]
        rows = [r for r in ws.iter_rows(min_row=3, values_only=True)
                if any(v is not None for v in r)]
        self.assertGreater(len(rows), 0)

    def test_compact_contains_potential_p_minus(self):
        """W998 lines (p;-) appear in compact sheet."""
        ws = self.wb["Report compact"]
        files = [ws.cell(row=r, column=1).value for r in range(3, ws.max_row + 1)]
        self.assertTrue(any(f and "module_a" in f for f in files))

    def test_extended_contains_p_n_false_positive(self):
        """W981 lines (p;n) produce 'false positive' auto-judgement in extended sheet."""
        ws = self.wb["Report extended"]
        headers = [c.value for c in ws[2]]
        aj_col = next(i for i, h in enumerate(headers) if h == "Auto judgement") + 1
        values = [ws.cell(row=r, column=aj_col).value for r in range(3, ws.max_row + 1)]
        self.assertTrue(any(v and "false positive" in v for v in values))

    def test_extended_contains_p_c_with_assembly_files(self):
        """W983 lines (p;c) with directory reference appear in extended sheet."""
        ws = self.wb["Report extended"]
        headers = [c.value for c in ws[2]]
        aj_col = next(i for i, h in enumerate(headers) if h == "Auto judgement") + 1
        values = [ws.cell(row=r, column=aj_col).value for r in range(3, ws.max_row + 1)]
        self.assertTrue(any(v and ".src." in v for v in values))

    def test_extended_p_c_unknown_shows_potential(self):
        """W983 without assembly dir info (p;c;?;?;?) shows 'Potential' judgement."""
        ws = self.wb["Report extended"]
        headers = [c.value for c in ws[2]]
        fn_col = next(i for i, h in enumerate(headers) if h == "File Name") + 1
        aj_col = next(i for i, h in enumerate(headers) if h == "Auto judgement") + 1
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=fn_col).value == "module_j.c":
                self.assertIn("Potential", ws.cell(row=row, column=aj_col).value)
                return
        self.skipTest("module_j.c row not found")

    def test_extended_deep_folder_path(self):
        """Detection from 5-level deep path (module_h.c) appears in extended sheet."""
        ws = self.wb["Report extended"]
        headers = [c.value for c in ws[2]]
        fn_col = next(i for i, h in enumerate(headers) if h == "File Name") + 1
        files = [ws.cell(row=r, column=fn_col).value for r in range(3, ws.max_row + 1)]
        self.assertIn("module_h.c", files)

    def test_extended_shallow_folder_path(self):
        """Detection from 1-level deep path (module_i.c) appears in extended sheet."""
        ws = self.wb["Report extended"]
        headers = [c.value for c in ws[2]]
        fn_col = next(i for i, h in enumerate(headers) if h == "File Name") + 1
        files = [ws.cell(row=r, column=fn_col).value for r in range(3, ws.max_row + 1)]
        self.assertIn("module_i.c", files)


@_SKIP
class TestIlConvVerbose(unittest.TestCase):
    """verbose=True exercises INFO print paths in parse.py and issuedb.py."""

    @classmethod
    def tearDownClass(cls):
        _remove_db()
        gc.collect()

    def test_verbose_prints_info_from_issuedb_and_parse(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        stem = tmp.name[:-5]
        try:
            output = _run_il_conv(
                ["-v", "--output", stem, str(TEST_LOG)],
                capture_stdout=True,
            )
            self.assertIn("INFO", output)
        finally:
            try:
                os.unlink(tmp.name)
            except FileNotFoundError:
                pass

    def test_verbose_mentions_log_file(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        stem = tmp.name[:-5]
        try:
            output = _run_il_conv(
                ["-v", "--output", stem, str(TEST_LOG)],
                capture_stdout=True,
            )
            self.assertIn("test.log", output)
        finally:
            try:
                os.unlink(tmp.name)
            except FileNotFoundError:
                pass


@_SKIP
class TestIlConvErrors(unittest.TestCase):
    """Error paths: missing files, wrong release note filename."""

    @classmethod
    def tearDownClass(cls):
        _remove_db()
        gc.collect()

    def test_nonexistent_relnotefile_raises_fnf(self):
        with patch("sys.argv", [
            "il_conv.py",
            "-x", str(XML_V63R1),
            "-r", "/nonexistent/readme_tricore_v6.3r1_inspector_v1.0r8.html",
            str(TEST_LOG),
        ]):
            with self.assertRaises(FileNotFoundError):
                il_conv()

    def test_relnotefile_name_not_starting_with_readme_tricore_raises(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".html", delete=False)
        tmp.close()
        try:
            with patch("sys.argv", [
                "il_conv.py",
                "-x", str(XML_V63R1),
                "-r", tmp.name,
                str(TEST_LOG),
            ]):
                with self.assertRaises(ValueError):
                    il_conv()
        finally:
            os.unlink(tmp.name)

    def test_relnotefile_unknown_compiler_version_raises(self):
        tmp_dir = tempfile.mkdtemp()
        rn_bad = Path(tmp_dir) / "readme_tricore_v9.9r9_inspector_v1.0r8.html"
        rn_bad.write_text("<html></html>")
        try:
            with patch("sys.argv", [
                "il_conv.py",
                "-x", str(XML_V63R1),
                "-r", str(rn_bad),
                str(TEST_LOG),
            ]):
                with self.assertRaises(ValueError):
                    il_conv()
        finally:
            rn_bad.unlink()
            Path(tmp_dir).rmdir()

    def test_nonexistent_xmlfile_raises_fnf(self):
        with patch("sys.argv", [
            "il_conv.py",
            "-x", "/nonexistent/issues.xml",
            "-r", str(RN_V63_V108),
            str(TEST_LOG),
        ]):
            with self.assertRaises(FileNotFoundError):
                il_conv()

    def test_multiple_logfiles_accepted(self):
        """il_conv accepts more than one logfile argument."""
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        stem = tmp.name[:-5]
        try:
            _run_il_conv(["--output", stem, str(TEST_LOG), str(TEST_LOG)])
        finally:
            try:
                os.unlink(tmp.name)
            except FileNotFoundError:
                pass


if __name__ == "__main__":
    unittest.main(verbosity=2)
